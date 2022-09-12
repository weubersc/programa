import pandas as pd
import os
import chardet
from listtocsv import listacsv as LCSV
import sqlite3
from programa.modulos.CONSULTA_SQL_DF import DF_SQL_DF as DF_SQL
from openpyxl import Workbook


class APUR():

    def __init__(self,ANO,SEM):
       print(ANO,SEM)
       self.dirpath=os.getcwd()
       self.ESTAB=(self.dirpath+"\\programa\\arquivos\\ESTAB.csv")
       self.PROD=(self.dirpath+"\\programa\\arquivos\\PROD.csv")
       self.PERG=(self.dirpath+"\\programa\\arquivos\\PERG.csv")
       self.wb=Workbook()
       DF_ESTAB=self.CSV_DF(self.ESTAB,"V")
#       print(DF_ESTAB.shape)
#       print(DF_ESTAB.dtypes)
       DF_ESTAB_ATUAL=DF_ESTAB[(DF_ESTAB["V1"]==ANO) &
                               (DF_ESTAB["V2"]==SEM) &
                                (DF_ESTAB["V23"]=="Ativo") &
                               (DF_ESTAB["V24"]==1)]
       
#       print(DF_ESTAB_ATUAL.shape)

       sql_query="select count(V3) as nestab,sum(V19) as CONV,"
       sql_query+="sum(V20) as GRAN,sum(V21) as SILO from TWEB_BR"

       DF_WEB_BR=DF_SQL(DF_ESTAB_ATUAL,sql_query,"TWEB_BR")

       lista_web_br=DF_WEB_BR.values.tolist()
       lista_webr=[]
       lista_webr.append("Brasil")
       for lwbr in lista_web_br[0]:
           lista_webr.append(lwbr)

#       print(lista_webr)

       
       sql_query="select V8,count(V3) as nestab,sum(V19) as CONV,"
       sql_query+="sum(V20) as GRAN,sum(V21) as SILO from TWEB"
       sql_query+=" group by V8"
       DF_WEB=DF_SQL(DF_ESTAB_ATUAL,sql_query,"TWEB")
       lista_web=DF_WEB.values.tolist()
       lista_web.insert(0,lista_webr)
       lista_web.insert(0,["UF","Q_WEB","CONV","GRAN","SILO"])
       self.add_sheet(lista_web,"QUEST_UF")
       LISTA_UF=DF_ESTAB_ATUAL["V8"].unique()
       for uf in LISTA_UF:
           DF_WEB=DF_ESTAB_ATUAL[(DF_ESTAB_ATUAL["V8"]==uf)]
           DF_WEB_R=DF_WEB[["V7","V3","V19","V20","V21"]]
           lista_web=DF_WEB_R.values.tolist()
           lista_web.insert(0,["UF","ESTAB","CONV","GRAN","SILO"])
           self.add_sheet(lista_web,uf)
       self.wb.remove(self.wb['Sheet'])
       self.wb.save(self.dirpath+'\\programa\\saida\\Quest_Web_UF_'+str(SEM)+'_'+str(ANO)+'.xlsx')
       
    def add_sheet(self,lista,var):
        self.wb.create_sheet(var)
        sheet=self.wb[var]
#        print(sheet.title)
        for index,lc in enumerate(lista):
            comp=len(lc)
            for y in range(0,comp):
                sheet.cell(row=index+1,column=y+1).value=lc[y]
       
       

                               

    def CSV_DF(self,arq,nvar):

        DB=pd.read_csv(arq,sep=";",
                                 encoding = "ISO-8859-1",header=None)
 
        tam=(DB.shape)
#        print(tam)
        temp=[]
        for i in range(1,tam[1]+1):
            temp.append(nvar+str(i))
#        print(temp)
        DB.columns = temp
        self.DBCOL=DB
        return(DB)

    

       
if __name__=="__main__":
    APUR(2021,1)
      
