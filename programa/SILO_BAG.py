import pandas as pd
import os
import chardet
from programa.modulos.listtocsv import listacsv as LCSV
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
import xlwt
from datetime import date
from codigos import Codigos as CODIG



class APURAÇÃO():
    def __init__(self,semestre,ano):
       data_atual = date.today()
       self.atual=str(data_atual.year)
       self.ano=str(ano)
       COD=CODIG()
       dic_cod=COD.ufcodigo
       self.semestre=semestre.upper()
       if self.semestre=="PRIMEIRO":
           self.sem=1
           data="30/06/"+self.ano
       if self.semestre=="SEGUNDO":
           self.sem=2
           data="31/12/"+self.ano
       print(semestre)
       print(self.sem)
       #print(self.ano,self.sem,sep="---")
       self.dirpath=os.getcwd()
       print(self.dirpath)
       self.ESTAB=(self.dirpath+"\\arquivos\\ESTAB.csv")
       self.PROD=(self.dirpath+"\\arquivos\\PROD.csv")
       self.PERG=(self.dirpath+"\\arquivos\\PERG.csv")
       
       self.alinharw=Alignment(horizontal='center',
                          vertical='center',
                              wrapText=True)
       self.alinhars=Alignment(horizontal='center',
                          vertical='center',
                              shrinkToFit=True)
       self.alinharj=Alignment(horizontal='center',
                          vertical='center',
                              justifyLastLine=True)
       
       self.alinhar1=Alignment(horizontal='center',
                          vertical='center')
       self.alinhar2=Alignment(horizontal='right',
                          vertical='center')
       self.Fonte=Font(name='Arial',
                     size=10,
                     bold=True)                     
       self.borda1=Border(top=Side(border_style='thin'),
                     bottom=Side(border_style='thin'))
       self.borda2=Border(top=Side(border_style='thin'),
                       bottom=Side(border_style='thin'),
                       left=Side(border_style='thin'),
                       right=Side(border_style='thin'))
       #print("OK")

       LISTA_REG_UF=[["11","12","13","14","15","16","17"],
                  ["21","22","23","24","25","26","27","28","29"],
                  ["31","32","33","35"],
                  ["41","42","43"],
                  ["50","51","52","53"]]
       
       self.LISTA_UF=["11","12","13","14","15","16","17",
                       "21","22","23","24","25","26","27","28","29",
                       "31","32","33","35",
                       "41","42","43",
                       "50","51","52","53"]

       DB_ESTAB=self.Ativo(self.ESTAB)
       #print(DB_ESTAB.shape)
       DB_PERG=self.Respondido(self.PERG)
       #print(DB_PERG.shape)
       dic_estab={}
       lista_cod_estab=[]
       for x,y in DB_ESTAB.iterrows():
           temp=[]
           temp.append([y["V6"],y["V7"],y["V8"]])
           dic_estab[y["V3"]]=temp
           
           

       resp="Houve produto do quadro acima armazenado em silos-bolsa, "
       resp+="na área do estabelecimento, em "+data+"?"
       resp2="Qual a totalidade de capacidade útil de silos bolsa, em quilogramas, "
       resp2+="utilizada com produtos do quadro acima, no estabelecimento, em "+data
       resp3="Qual o produto armazenado em silo bolsa?"
       print(resp)
       print(resp2)
       perguntas=DB_PERG["V5"].unique()
       for pp in perguntas:
           print(pp)
       DB_P5=DB_PERG[((DB_PERG["V5"]==resp) &
                     (DB_PERG["V6"]=="Sim") )]
       print(DB_P5.shape)
       dic_p5={}
       for x,y in DB_P5.iterrows():
           lista_cod_estab.append(y["V3"])
           dic_p5[y["V3"]]=True



       LUF=self.UF(DB_P5,dic_estab)


       DB_P5_2=DB_PERG[((DB_PERG["V5"]==resp2))]
       #print(DB_P5_2.shape)
       l_quant=[]
       dic_quant={}
       for x,y in DB_P5_2.iterrows():
           lista_cod_estab.append(y["V3"])
           try:
               temp=[]
               var=dic_p5[y["V3"]]
               ll=dic_estab[y["V3"]]
               temp=ll[0]+[y["V6"]]
               l_quant.append(temp)
               dic_quant[y["V3"]]=y["V6"]
           except:
               pass
            
       LUFQ=self.UFQ(l_quant)
       DB_P5_3=DB_PERG[((DB_PERG["V5"]==resp3))]
       #print(DB_P5_3.shape)
       l_prod=[]
       for x,y in DB_P5_3.iterrows():
           lista_cod_estab.append(y["V3"])
           try:
               temp=[]
               quant=dic_quant[y["V3"]]
               ll=dic_estab[y["V3"]]
               temp=ll[0]+[quant,y["V6"]]
               l_prod.append(temp)
           except:
               pass
            
##
##       for ll in l_prod:
##           print(ll)
  
       LUFP=self.UFP(l_prod)

       self.analise(lista_cod_estab,DB_PERG,resp,resp2,resp3)
       lista_final=[]
       for lup in LUFP:
           print(lup)
           coduf=(dic_cod[lup[0]])
           ninf=lup[1]+lup[2]+lup[3]
           lista_final.append([coduf,ninf,lup[1],lup[2],
                         lup[3],lup[7],lup[4],lup[5],
                         lup[6]])


       LCSV(lista_final,"silo.csv",";")
       self.excel(lista_final)
           


              
    def espaco(self,valor):
        vf=("{:,}".format(valor).replace(",", " "))
        return(vf)

    def percentual(self,som1,som2):
        per=[0,0,0,0]
        for k in range(0,3):
            temp=round(som2[k]/som1[k]*100,1)
            per[k]=str(temp)
            self.perc[k]=per[k].replace(".",",")
    def percent(self,som1,som2):
        temp=round(som2/som1*100,1)
        per=str(temp)
        per=per.replace(".",",")
        return(per)

    def Ativo(self,arq):

        DB=pd.read_csv(arq,sep=";",
                                 encoding = "ISO-8859-1",header=None)
 
        tam=(DB.shape)
        #print(tam)
        temp=[]
        for i in range(1,tam[1]+1):
            temp.append("V"+str(i))
        #print(temp)
        DB.columns = temp
        self.DBCOL=DB

##        for col in self.DBCOL.columns: 
##                print(col)

        DB_ATIVO=self.DBCOL[((self.DBCOL['V1']==int(self.ano)) &
                         (self.DBCOL['V2']==self.sem) & 
                         (self.DBCOL['V23']=='Ativo'))]
        #print("ANO-ATUAL")
       # print(DB_ATIVO.shape)

        return(DB_ATIVO)

    def Respondido(self,arq):
        DB=pd.read_csv(arq,sep=";",
                                 encoding = "ISO-8859-1",header=None)
 
        tam=(DB.shape)
        #print(tam)
        temp=[]
        for i in range(1,tam[1]+1):
            temp.append("V"+str(i))
        #print(temp)
        DB.columns = temp
##        for col in DB.columns: 
##                print(col)
        DB_RESPONDIDO=DB[((DB['V1']==int(self.ano)) &
                         (DB['V2']==self.sem))]
        #print("PERGUNTAS")
        #print(DB_RESPONDIDO.shape)


        return(DB_RESPONDIDO)

    def UF(self,DB,d_estab):
        lista_uf=[]
        contbr=0
        for uf in self.LISTA_UF:
            cuf=int(uf)
            #print(uf)
            cont=0
            for x,y in DB.iterrows():
                lista=d_estab[y["V3"]]
                luf=(lista[0])
                if luf[0]==cuf:
                    cont+=1
                    contbr+=1
            if cont>0:
                lista_uf.append([uf,cont])

        lista_uf.append(["BR",contbr])
        return(lista_uf)
        
    def UFQ(self,listaq):
        lista_uf=[]
        contbr=0
        somabr=0
        for uf in self.LISTA_UF:
            cuf=int(uf)
            #print(uf)
            soma=0
            cont=0
            for lq in listaq:
                luf=(lq[0])
                if luf==cuf:
                    cont+=1
                    contbr+=1
                    soma+=int(lq[3])
                    somabr+=int(lq[3])
            if cont>0:
                #print(uf,cont,soma)
                lista_uf.append([uf,soma])

        #print("BR",contbr,somabr)
        lista_uf.append(["BR",somabr])
        return(lista_uf)

    def UFP(self,listap):
        lista_uf=[]
        contbr=[0,0,0]
        somabr=[0,0,0]
        for uf in self.LISTA_UF:
            cuf=int(uf)
            #print(uf)
            soma=[0,0,0]
            cont=[0,0,0]
            for lq in listap:
                luf=(lq[0])
                var=int(lq[4])-1
                if luf==cuf:
                    cont[var]+=1
                    contbr[var]+=1
                    soma[var]+=int(lq[3])
                    somabr[var]+=int(lq[3])
            if cont[0]>0 or cont[1]>0 or cont[2]>0 :
                temp=[]
                somat=soma[0]+soma[1]+soma[2]
                #print(uf,cont,soma,somat)
                temp=[uf]+cont+soma+[somat]
                lista_uf.append(temp)
        somat=somabr[0]+somabr[1]+somabr[2]
        #print("BR",contbr,somabr,somat)
        tempbr=["00"]+contbr+somabr+[somat]
        lista_uf.append(tempbr)
        return(lista_uf)

    def analise(self,lista,db,r1,r2,r3):
        #print(len(lista))
        lista_dif=list(set(lista))
        #print(len(lista_dif))
        l_anal=[]
        for ll in lista_dif:
            #print(ll)
            teste=[False,False,False]
            temp=[0,0,0]
            DBC=db[(db["V3"]==ll)]
            for x,y in DBC.iterrows():
                if y["V5"]==r1:
                        teste[0]=True
                        temp[0]=y["V6"]
                if y["V5"]==r2:
                        teste[1]=True
                        temp[1]=y["V6"]
                if y["V5"]==r3:
                        teste[2]=True
                        temp[2]=y["V6"]
            temp.insert(0,ll)
            tt=temp+teste
            l_anal.append(tt)
##        for la in l_anal:
##            if la[4]==False or la[5]==False or la[6]==False or la[1]=="Nao" or la[1]=="Não":
##                #print(la)
                        
    def excel(self,lista_excel):
        wb=Workbook()
        if self.sem==1:
            data_ref="30/06/"+(self.ano)
        elif self.sem==2:
            data_ref="31/12/"+(self.ano)
        linha1="Número de Estabelecimentos e quantidade em (kg) de produto armazenado em silo-bolsa na \n"
        linha1+="área do estabelecimento, em "+ data_ref+" em nível de unidade da federeção e Brasil."
        wb.create_sheet("Silo bag")
        sheet=wb["Silo bag"]
        print(sheet.title)
        sheet.merge_cells("A1:I2")
        sheet.merge_cells("A3:A4")
        sheet.merge_cells("B3:E3")
        sheet.merge_cells("F3:I3")
        sheet.cell(row=1,column=1).value=linha1
        sheet.cell(row=1,column=1).font=self.Fonte
        sheet.row_dimensions[1].height = 30
        sheet.cell(row=1,column=1).alignment=self.alinharw
        sheet.cell(row=3,column=1).value="Unidade da Federação"
        sheet.cell(row=3,column=1).font=self.Fonte
        sheet.column_dimensions["A"].width = len("Unidade da Federação")+5
        sheet.cell(row=3,column=2).value="Nº de estabelecimentos"
        sheet.cell(row=3,column=2).font=self.Fonte
        sheet.cell(row=3,column=6).value="Quantidade (Kg)"
        sheet.cell(row=3,column=6).font=self.Fonte
        
        lista_rot=["Total","Soja","Milho","Outros"]
        for j in range(2,10):
           if j>5:
                sheet.cell(row=4,column=j).value=lista_rot[j-6]
                sheet.cell(row=4,column=j).font=self.Fonte
           else:
                sheet.cell(row=4,column=j).value=lista_rot[j-2]
                sheet.cell(row=4,column=j).font=self.Fonte
        valor=0                
        for ind,ll in enumerate(lista_excel):
            sheet.cell(row=5+ind,column=1).value=ll[0]
            sheet.cell(row=5+ind,column=1).font=self.Fonte
            for mm in range(1,len(ll)):
                 vv=self.espaco(ll[mm])
                 sheet.cell(row=5+ind,column=mm+1).value=vv
                 sheet.cell(row=5+ind,column=mm+1).alignment=self.alinhar1
                 if len(vv)>valor:
                     valor=len(vv)

        letra=["A","B","C","D","E","F","G","H","I"]
        for ind in range(6,10):
            sheet.column_dimensions[letra[ind-1]].width=valor+2
            

                 

        for le in letra:
            for inf in range(1,5):
                if inf==1 or inf==2:
                    sheet[le+str(inf)].border=self.borda1
                elif inf>2 and le!="A" and le!="I":
                    sheet[le+str(inf)].border=self.borda2
                    sheet[le+str(inf)].alignment=self.alinhar1
                else:
                    sheet[le+str(inf)].border=self.borda1
                    sheet[le+str(inf)].alignment=self.alinhar1
        wb.remove(wb['Sheet'])
        nome="silo_bag_"+self.semestre+"_SEMESTRE_"+str(self.ano)
        print(self.dirpath)
        wb.save(self.dirpath+"\\saida\\"+nome+".xlsx")


    




                 
       
if __name__=="__main__":
    APURAÇÃO("SEGUNDO",2021)       
      
