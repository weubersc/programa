import pandas as pd
import os
import chardet
from listtocsv import listacsv as LCSV
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from openpyxl.styles import Border, Side, Font
import xlwt
from datetime import date



class APURAÇÃO():
    def __init__(self,semestre,ano,mes_apu):
       data_atual = date.today()
       self.atual=str(data_atual.year)
       self.ano=str(ano)
       self.ant=str(int(self.ano)-1)
       if semestre=="PRIMEIRO":
           self.sem=1
       if semestre=="SEGUNDO":
           self.sem=2
       self.apur=mes_apu
       print(self.ano,self.ant,self.sem,sep="---")
       self.dirpath=os.getcwd()
       self.ESTAB=(self.dirpath+"\\programa\\arquivos\\ESTAB.csv")
       self.PROD=(self.dirpath+"\\programa\\arquivos\\PROD.csv")
       self.PERG=(self.dirpath+"\\programa\\arquivos\\PERG.csv")
       print("OK")
       LISTA_PROD5=['Algodão (em pluma)',
                 'Algodão (em caroço)',
                 'Caroço de Algodão',
                 'Semente de Algodão',
                 'Arroz (em casca)',
                 'Arroz Beneficiado',
                 'Semente de Arroz',
                 'Café Arábica (em grão)',
                 'Café Canephora (em grão)',
                 'Feijão Preto (em grão)',
                 'Feijão de Cor (em grão)',
                 'Milho (em grão)',
                 'Semente de Milho',
                 'Soja (em grão)',
                 'Semente de Soja',
                 'Trigo (em grão)',
                 'Semente de Trigo',
                 'Outros Grãos e Sementes'
                 ]

       LISTA_REG_UF=[["11","12","13","14","15","16","17"],
                  ["21","22","23","24","25","26","27","28","29"],
                  ["31","32","33","35"],
                  ["41","42","43"],
                  ["50","51","52","53"]]

       lista_ativo=self.Ativo(self.ESTAB)
       lista_ativo_ant=self.Ativo_ant(self.ESTAB)
       dic_respondido=self.Respondido(self.PERG)

       contador_t=0
       contador_ant=0
       somat=[0,0,0]
       somant=[0,0,0]
       somar=[0,0,0]
       somantr=[0,0,0]
       for ll in lista_ativo:
           somat[0]+=int(ll[18])
           somat[1]+=int(ll[19])
           somat[2]+=int(ll[20])
           try:
               if dic_respondido[ll[2]]==True:
                   somar[0]+=int(ll[18])
                   somar[1]+=int(ll[19])
                   somar[2]+=int(ll[20])
                   contador_t+=1
           except:
               pass
       for ll in lista_ativo_ant:
           somant[0]+=int(ll[18])
           somant[1]+=int(ll[19])
           somant[2]+=int(ll[20])
           try:
               if dic_respondido[ll[2]]==True:
                   somantr[0]+=int(ll[18])
                   somantr[1]+=int(ll[19])
                   somantr[2]+=int(ll[20])
                   contador_ant+=1
           except:
               pass
       print(contador_t)
       print(len(lista_ativo),contador_t)
       per_inf=self.percent(len(lista_ativo),contador_t)
       print(somat[0],somar[0])
       print(somat[1],somar[1])
       print(somat[2],somar[2])
       print(contador_ant)
       self.perc=["0","0","0"]
       self.percentual(somat,somar)
       print(self.perc)
       print(len(lista_ativo_ant),contador_ant)
       print(somant[0],somantr[0])
       print(somant[1],somantr[1])
       print(somant[2],somantr[2])

       "Saída para o Excel - CAPACIDADE"

       lista_rot=["Nº de estabelecimentos",
                  "Convencial",
                  "Granelizados",
                  "Silo"]
       lista_cab=["Variável","Total","Apurado", "%(Apurado/total)"]

       arquivo=Workbook()
       folha=arquivo.active
       folha.title="Apuração provisória"
       for j in range(1,5):
           folha.cell(row=1,column=j).font=Font(bold=True)
           folha.cell(row=1,column=j).value=lista_cab[j-1]
       folha.cell(row=2,column=1).value=lista_rot[0]

       vf=self.espaco(len(lista_ativo))
       folha.cell(row=2,column=2).value=vf
       vf=self.espaco(contador_t)
       folha.cell(row=2,column=3).value=vf
       folha.cell(row=2,column=4).value=per_inf
       for k in range(3,6):
           folha.cell(row=k,column=1).value=lista_rot[k-2]
           vf=self.espaco(somat[k-3])
           folha.cell(row=k,column=2).value=vf
           vf=self.espaco(somar[k-3])
           folha.cell(row=k,column=3).value=vf
           folha.cell(row=k,column=4).value=self.perc[k-3]
       folha.merge_cells("A6:D6")
       texto_apur="*Apuração provisória realizada em "+self.apur+ " de "+self.atual
       folha.cell(row=6,column=1).value=texto_apur
       fontStyle = Font(size = "8",italic=True)
       folha.cell(row=6,column=1).font=fontStyle
       for k in range(2,6):
           for j in range(2,5):
               folha.cell(row=k,column=j).alignment=Alignment(horizontal='center',vertical='center')
       #arquivo.save("rf.xlsx")

##       arquivo=xlwt.Workbook(encoding="utf-8")
##       folha1=arquivo.add_sheet("Apuração provisória")
##       row=folha1.row(0)
##       for m in range(0,4):
##           row.write(m,lista_cab[m])
##       for j in range(1,4):
##           row=folha1.row(j)
##           row.write(0,lista_rot[j-1])
##           row.write(1,somat[j-1])
##           row.write(2,somar[j-1])
##           row.write(3,self.perc[j-1])
##
##       arquivo.save("teste.xls")
##       print("excel criado")

       
       self.comp=[]
       lista_prod=self.Prod(self.PROD)
       lista_prod_ant=self.Prod_ant(self.PROD)
       self.comparacao(LISTA_PROD5,lista_prod,lista_prod_ant,
                       dic_respondido)
       l_prod=["Algodão (em pluma)",
                "Algodão (em caroço)",
                "Caroço de Algodão",
                "Arroz (em casca)",
                "Arroz Beneficiado",
                "Café Arábica (em grão)",
                "Café Canephora (em grão)",
                "Feijão Preto (em grão)",
                "Feijão de Cor (em grão)",
                "Milho (em grão)",
                "Soja (em grão)",
                "Trigo (em grão)"]
       
       lista_prod=[]
       for ll in l_prod:
           for comp in self.comp:
               if comp[0]==ll:
                   vf1=self.espaco(comp[2])
                   vf2=self.espaco(comp[3])
                   lista_prod.append([comp[0],comp[1],vf1,vf2,comp[4]])
       #arquivo=load_workbook('rf.xlsx')
       arquivo.create_sheet("Comparação Produto")
       folha2=arquivo["Comparação Produto"]
       folha2.merge_cells("A1:A2")
       folha2.merge_cells("B1:B2")
       folha2.merge_cells("C1:D1")
       celu=folha2.cell(row=1,column=1)
       celu.value="Produto"
       celu.alignment=Alignment(horizontal='center',vertical='center')
       celu.font = Font(bold=True)
       celu=folha2.cell(row=1,column=2)
       celu.value="Nº Inf."
       celu.alignment=Alignment(horizontal='center',vertical='center')
       celu.font = Font(bold=True)
       celu=folha2.cell(row=1,column=3)
       celu.value="Capacidade (kg)"
       celu.alignment=Alignment(horizontal='center',vertical='center')
       celu.font = Font(bold=True)
       folha2.cell(row=1,column=5).value="Percentual(%)"
       folha2.cell(row=1,column=5).font = Font(bold=True)
       texto1=str(self.sem)+"º sem. "+self.ano
       folha2.cell(row=2,column=3).value=texto1
       folha2.cell(row=2,column=3).font = Font(bold=True)
       texto2=str(self.sem)+"º sem. "+self.ant
       folha2.cell(row=2,column=4).value=texto2
       folha2.cell(row=2,column=4).font = Font(bold=True)
       texto3=self.ano+"/"+self.ant
       folha2.cell(row=2,column=5).value=texto3
       folha2.cell(row=2,column=5).font = Font(bold=True)
       folha2.cell(row=15,column=1).font=fontStyle
       folha2.cell(row=15,column=1).value=texto_apur
       #arquivo.save("rf.xlsx")
       for k in range(3,15):
           for j in range(2,6):
               folha2.cell(row=k,column=j).alignment=Alignment(horizontal='center',vertical='center')
       for index,ll in enumerate(lista_prod):
           for k in range(1,6):
               folha2.cell(row=index+3,column=k).value=ll[k-1]
       arquivo.save(self.dirpath+"\\programa\\saida\\apuração provisória_"+
                    str(data_atual)+'.xlsx')

    def espaco(self,valor):
        vf=("{:,}".format(valor).replace(",", " "))
        return(vf)

    def percentual(self,som1,som2):
        per=[0,0,0,0]
        for k in range(0,3):
            temp=round(som2[k]/som1[k]*100,1)
            per[k]=str(temp)
            self.perc[k]=per[k].replace(".",",")
    def percent(self,som1,som2):
        temp=round(som2/som1*100,1)
        per=str(temp)
        per=per.replace(".",",")
        return(per)

    def Ativo(self,arq):

        DB=pd.read_csv(arq,sep=";",
                                 encoding = "ISO-8859-1",header=None)
 
        tam=(DB.shape)
        #print(tam)
        temp=[]
        for i in range(1,tam[1]+1):
            temp.append("V"+str(i))
        #print(temp)
        DB.columns = temp
        self.DBCOL=DB

        for col in self.DBCOL.columns: 
                print(col)

        DB_ATIVO=self.DBCOL[((self.DBCOL['V1']==int(self.ano)) &
                         (self.DBCOL['V2']==self.sem) & 
                         (self.DBCOL['V23']=='Ativo'))]
        print("ANO-ATUAL")
        print(DB_ATIVO.shape)

        LISTA_ATIVO=DB_ATIVO.values.tolist()
        return(LISTA_ATIVO)

    def Respondido(self,arq):
        DB=pd.read_csv(arq,sep=";",
                                 encoding = "ISO-8859-1",header=None)
 
        tam=(DB.shape)
        #print(tam)
        temp=[]
        for i in range(1,tam[1]+1):
            temp.append("V"+str(i))
        #print(temp)
        DB.columns = temp
        for col in DB.columns: 
                print(col)
        DB_RESPONDIDO=DB[((DB['V1']==int(self.ano)) &
                         (DB['V2']==self.sem) &
                         (DB['V4']==1))]
        print("RESPONDIDO")
        print(DB_RESPONDIDO.shape)

        LISTA_PROV=DB_RESPONDIDO.values.tolist()
        dic_resp={}
        for ll in LISTA_PROV:
            dic_resp[ll[2]]=True

        return(dic_resp)

    def Ativo_ant(self,arq):

        DB=pd.read_csv(arq,sep=";",
                                 encoding = "ISO-8859-1",header=None)
 
        tam=(DB.shape)
        #print(tam)
        temp=[]
        for i in range(1,tam[1]+1):
            temp.append("V"+str(i))
        #print(temp)
        DB.columns = temp
        self.DBCOL=DB

        for col in self.DBCOL.columns: 
                print(col)

        DB_ATIVO=self.DBCOL[((self.DBCOL['V1']==int(self.ant)) &
                         (self.DBCOL['V2']==self.sem) & 
                         (self.DBCOL['V23']=='Ativo'))]
        print("ANO=ANTERIOR")
        print(DB_ATIVO.shape)

        LISTA_ATIVO_ANT=DB_ATIVO.values.tolist()
        return(LISTA_ATIVO_ANT)

    def Prod(self,arq):

        DB=pd.read_csv(arq,sep=";",
                                 encoding = "ISO-8859-1",header=None)
 
        tam=(DB.shape)
        print(tam)
        temp=[]
        for i in range(1,tam[1]+1):
            temp.append("V"+str(i))
        print(temp)
        DB.columns = temp
        self.DBCOL=DB

        for col in self.DBCOL.columns: 
                print(col)

        DB_PROD=self.DBCOL[((self.DBCOL['V1']==int(self.ano)) &
                         (self.DBCOL['V2']==self.sem))]

        print(DB_PROD.shape)
        LISTA_PROD=DB_PROD.values.tolist()
        return(LISTA_PROD)

    def Prod_ant(self,arq):

        DB=pd.read_csv(arq,sep=";",
                                 encoding = "ISO-8859-1",header=None)
 
        tam=(DB.shape)
        print(tam)
        temp=[]
        for i in range(1,tam[1]+1):
            temp.append("V"+str(i))
        print(temp)
        DB.columns = temp
        self.DBCOL=DB

        for col in self.DBCOL.columns: 
                print(col)

        DB_PROD=self.DBCOL[((self.DBCOL['V1']==int(self.ant)) &
                         (self.DBCOL['V2']==self.sem))]

        print(DB_PROD.shape)
        LISTA_PROD=DB_PROD.values.tolist()
        return(LISTA_PROD)
    
    def comparacao(self,listap,lista1,lista2,dicr):
        for lp in listap:
            soma1=0
            contador1=0
            soma2=0
            contador2=0
            L1=[]
            L2=[]
            for l1 in lista1:
                if l1[3]==lp:
                   L1.append([l1[2],l1[4]])
                   soma1+=int(l1[4])
                   contador1+=1
            for l2 in lista2:
                if l2[3]==lp:
                       L2.append([l2[2],l2[4]])
                       soma2+=int(l2[4])
                       contador2+=1
            tot1=0
            tot2=0
            cont=0
            for m1 in L1:
                for m2 in L2:
                    if m1[0]==m2[0]:
                        tot1+=m1[1]
                        tot2+=m2[1]
                        cont+=1

            print(lp,contador1,contador2,soma1,soma2,sep=";")
            print(lp,cont,tot1,tot2,sep=";")
            try:
                temp=round((tot1-tot2)/tot2*100,1)
                pp=str(temp)
                pp=pp.replace(".",",")
            except:
                pp="0,0"
            self.comp.append([lp,cont,tot1,tot2,pp])
            print("-------------------")                 
       
     
      
